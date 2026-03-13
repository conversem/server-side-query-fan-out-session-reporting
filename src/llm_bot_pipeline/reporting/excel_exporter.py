"""
Excel exporter -- reads from any StorageBackend and generates .xlsx reports.

Works with both SQLite and BigQuery backends, making it possible to
export reports from any processing mode.

Requires: openpyxl (already in requirements.txt)
"""

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Sequence, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config.constants import (
    AUTOFIT_SAMPLE_ROWS,
    EXCEL_DATE_FORMAT,
    EXCEL_DATETIME_FORMAT,
    TABLE_CLEAN_BOT_REQUESTS,
    TABLE_DAILY_SUMMARY,
    TABLE_QUERY_FANOUT_SESSIONS,
    TABLE_URL_PERFORMANCE,
    VALID_DATE_COLUMNS,
    VALID_TABLE_NAMES,
)
from ..storage import StorageBackend, StorageError

logger = logging.getLogger(__name__)

_VALID_COLUMN_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]*$")

# Default tables included in a full report
DEFAULT_REPORT_TABLES = (
    TABLE_DAILY_SUMMARY,
    TABLE_URL_PERFORMANCE,
    TABLE_CLEAN_BOT_REQUESTS,
    TABLE_QUERY_FANOUT_SESSIONS,
)

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


class ExcelExporter:
    """Export data from any StorageBackend to .xlsx workbooks.

    Usage::

        from llm_bot_pipeline.storage import get_backend
        from llm_bot_pipeline.reporting.excel_exporter import ExcelExporter

        backend = get_backend("sqlite", db_path="data/llm-bot-logs.db")
        backend.initialize()

        exporter = ExcelExporter(backend)
        exporter.export_report("report.xlsx", start_date=date(2026, 1, 1))
    """

    def __init__(self, backend: StorageBackend) -> None:
        self._backend = backend

    def export_table(
        self,
        table_name: str,
        output_path: Union[str, Path],
        filters: Optional[dict] = None,
        limit: Optional[int] = None,
        sheet_name: Optional[str] = None,
    ) -> Path:
        """Export a single table to an xlsx file.

        Args:
            table_name: Name of the table to export
            output_path: Destination file path (.xlsx)
            filters: Optional dict of column=value equality filters
            limit: Optional row limit
            sheet_name: Sheet name (defaults to table_name)

        Returns:
            Path to the written file
        """
        output_path = Path(output_path)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name or table_name

        sql, params = self._build_select(table_name, filters, limit)
        rows = self._query(sql, params)

        if rows:
            self._write_sheet(ws, rows)
        else:
            ws.append(["No data"])

        wb.save(str(output_path))
        logger.info("Exported %s (%d rows) to %s", table_name, len(rows), output_path)
        return output_path

    def export_report(
        self,
        output_path: Union[str, Path],
        tables: Optional[Sequence[str]] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> Path:
        """Export a multi-sheet workbook with several tables.

        Args:
            output_path: Destination file path (.xlsx)
            tables: List of table names to include (defaults to standard set)
            start_date: Optional date filter (lower bound)
            end_date: Optional date filter (upper bound)

        Returns:
            Path to the written file
        """
        output_path = Path(output_path)
        tables = tables or DEFAULT_REPORT_TABLES
        wb = Workbook()

        for idx, table in enumerate(tables):
            if not self._backend.table_exists(table):
                logger.warning("Table %s does not exist, skipping", table)
                continue

            if idx == 0:
                ws = wb.active
                ws.title = table
            else:
                ws = wb.create_sheet(title=table)

            date_col = self._detect_date_column(table)
            filters = {}
            if date_col and start_date:
                filters["__date_col"] = date_col
                filters["__start_date"] = start_date.isoformat()
                if end_date:
                    filters["__end_date"] = end_date.isoformat()

            sql, params = self._build_select(table, filters)
            rows = self._query(sql, params)

            if rows:
                self._write_sheet(ws, rows)
            else:
                ws.append(["No data"])

            logger.info("  Sheet '%s': %d rows", table, len(rows))

        # Remove default empty sheet if we created named ones
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(str(output_path))
        logger.info("Report saved to %s (%d sheets)", output_path, len(wb.sheetnames))
        return output_path

    def export_query(
        self,
        sql: str,
        output_path: Union[str, Path],
        sheet_name: str = "Results",
    ) -> Path:
        """Export results of an arbitrary SQL query to xlsx.

        Args:
            sql: SQL query to execute
            output_path: Destination file path
            sheet_name: Name for the worksheet

        Returns:
            Path to the written file
        """
        output_path = Path(output_path)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        rows = self._query(sql)

        if rows:
            self._write_sheet(ws, rows)
        else:
            ws.append(["No data"])

        wb.save(str(output_path))
        logger.info("Query export: %d rows to %s", len(rows), output_path)
        return output_path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _query(self, sql: str, params: Optional[dict] = None) -> list[dict]:
        try:
            return self._backend.query(sql, params)
        except StorageError as e:
            logger.error("Query failed: %s", e)
            return []

    def _build_select(
        self,
        table: str,
        filters: Optional[dict] = None,
        limit: Optional[int] = None,
    ) -> tuple[str, dict]:
        """Build a parameterized SELECT statement with validation.

        Returns:
            Tuple of (sql_string, params_dict) for safe parameterized execution.

        Raises:
            ValueError: If table name or column names fail validation.
        """
        if table not in VALID_TABLE_NAMES:
            raise ValueError(
                f"Invalid table name: {table!r}. "
                f"Must be one of: {sorted(VALID_TABLE_NAMES)}"
            )

        clauses: list[str] = []
        params: dict = {}

        if filters:
            date_col = filters.pop("__date_col", None)
            start = filters.pop("__start_date", None)
            end = filters.pop("__end_date", None)

            if date_col:
                if date_col not in VALID_DATE_COLUMNS:
                    raise ValueError(
                        f"Invalid date column: {date_col!r}. "
                        f"Must be one of: {sorted(VALID_DATE_COLUMNS)}"
                    )
                if start:
                    clauses.append(f"{date_col} >= :start_date")
                    params["start_date"] = start
                    if end:
                        clauses.append(f"{date_col} <= :end_date")
                        params["end_date"] = end

            for col, val in filters.items():
                if not _VALID_COLUMN_RE.match(col):
                    raise ValueError(
                        f"Invalid column name: {col!r}. "
                        "Column names must contain only letters, digits, "
                        "and underscores."
                    )
                param_key = f"filter_{col}"
                clauses.append(f"{col} = :{param_key}")
                params[param_key] = str(val)

        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        limit_clause = f" LIMIT {int(limit)}" if limit else ""
        return f"SELECT * FROM {table}{where}{limit_clause}", params

    @staticmethod
    def _detect_date_column(table: str) -> Optional[str]:
        """Return the likely date-filtering column for a given table."""
        mapping = {
            TABLE_DAILY_SUMMARY: "request_date",
            TABLE_URL_PERFORMANCE: "request_date",
            TABLE_CLEAN_BOT_REQUESTS: "request_date",
            TABLE_QUERY_FANOUT_SESSIONS: "session_date",
        }
        return mapping.get(table)

    @staticmethod
    def _write_sheet(ws, rows: list[dict]) -> None:
        """Write rows to a worksheet with headers, formatting, and auto-width."""
        if not rows:
            return

        headers = list(rows[0].keys())
        ExcelExporter._write_header_row(ws, headers)
        ExcelExporter._write_data_rows(ws, rows, headers)
        ExcelExporter._auto_fit_columns(ws, headers, len(rows))
        ws.freeze_panes = "A2"

    @staticmethod
    def _write_header_row(ws, headers: list[str]) -> None:
        """Style and write the header row."""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT

    @staticmethod
    def _write_data_rows(ws, rows: list[dict], headers: list[str]) -> None:
        """Write data cells with date formatting."""
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row.get(header)
                if isinstance(val, datetime):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.number_format = EXCEL_DATETIME_FORMAT
                elif isinstance(val, date):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.number_format = EXCEL_DATE_FORMAT
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)

    @staticmethod
    def _auto_fit_columns(ws, headers: list[str], row_count: int) -> None:
        """Auto-fit column widths based on content."""
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(row_count + 2, AUTOFIT_SAMPLE_ROWS + 2)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

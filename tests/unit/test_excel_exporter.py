"""Tests for ExcelExporter."""

import sqlite3
import tempfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from llm_bot_pipeline.config.constants import TABLE_DATA_FRESHNESS
from llm_bot_pipeline.reporting.excel_exporter import ExcelExporter
from llm_bot_pipeline.storage import get_backend


@pytest.fixture()
def backend_with_data(tmp_path):
    """SQLite backend pre-loaded with sample data."""
    db_path = tmp_path / "test.db"
    backend = get_backend("sqlite", db_path=db_path)
    backend.initialize()

    conn = sqlite3.connect(str(db_path))
    for i in range(5):
        d = f"2026-01-{10 + i:02d}"
        conn.execute(
            """INSERT INTO daily_summary
               (request_date, bot_provider, bot_name, bot_category,
                total_requests, unique_urls, unique_hosts,
                successful_requests, error_requests,
                redirect_requests, _aggregated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                d,
                "OpenAI",
                "GPTBot",
                "training",
                100 + i * 10,
                5 + i,
                1,
                90,
                5,
                5,
                "2026-01-20T00:00:00",
            ),
        )
        conn.execute(
            """INSERT INTO url_performance
               (request_date, request_host, url_path,
                total_bot_requests, unique_bot_providers, unique_bot_names,
                training_hits, user_request_hits,
                successful_requests, error_requests,
                first_seen, last_seen, _aggregated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                d,
                "example.com",
                f"/page/{i}",
                50 + i,
                1,
                1,
                40,
                10,
                45,
                5,
                f"{d}T00:00:00",
                f"{d}T23:59:59",
                "2026-01-20T00:00:00",
            ),
        )
    conn.commit()
    conn.close()

    yield backend
    backend.close()


class TestExportTable:
    def test_export_table(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "table.xlsx"
        exporter.export_table("daily_summary", out)

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.title == "daily_summary"
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "request_date" in headers
        assert ws.max_row == 6  # 1 header + 5 data


class TestExportReport:
    def test_export_report_multi_sheet(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "report.xlsx"
        exporter.export_report(out, tables=["daily_summary", "url_performance"])

        wb = load_workbook(str(out))
        assert "daily_summary" in wb.sheetnames
        assert "url_performance" in wb.sheetnames

    def test_missing_table_skipped(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "report.xlsx"
        exporter.export_report(out, tables=["daily_summary", "nonexistent_table"])

        wb = load_workbook(str(out))
        assert "nonexistent_table" not in wb.sheetnames
        assert "daily_summary" in wb.sheetnames


class TestExportQuery:
    def test_export_query(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "query.xlsx"
        exporter.export_query(
            "SELECT * FROM daily_summary WHERE total_requests > 110", out
        )

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.max_row > 1

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "request_date" in headers
        assert "bot_provider" in headers
        assert "total_requests" in headers

        col_request_date = headers.index("request_date") + 1
        col_bot_provider = headers.index("bot_provider") + 1
        col_total_requests = headers.index("total_requests") + 1

        first_data_row = 2
        first_request_date = ws.cell(row=first_data_row, column=col_request_date).value
        first_bot_provider = ws.cell(row=first_data_row, column=col_bot_provider).value
        first_total_requests = ws.cell(
            row=first_data_row, column=col_total_requests
        ).value

        assert str(first_request_date) == "2026-01-12"
        assert first_bot_provider == "OpenAI"
        assert first_total_requests == 120
        assert isinstance(first_total_requests, (int, float))

    def test_export_query_different_filter_returns_different_values(
        self, backend_with_data, tmp_path
    ):
        """Different query returns different content; assertions would fail if wrong."""
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "query2.xlsx"
        exporter.export_query(
            "SELECT * FROM daily_summary WHERE total_requests <= 110 ORDER BY total_requests",
            out,
        )
        wb = load_workbook(str(out))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_total_requests = headers.index("total_requests") + 1
        first_total = ws.cell(row=2, column=col_total_requests).value
        assert first_total == 100
        assert isinstance(first_total, (int, float))


class TestDateFilter:
    def test_date_filter(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "filtered.xlsx"
        exporter.export_report(
            out,
            tables=["daily_summary"],
            start_date=date(2026, 1, 12),
            end_date=date(2026, 1, 13),
        )

        wb = load_workbook(str(out))
        ws = wb.active
        data_rows = ws.max_row - 1
        assert data_rows == 2


class TestEmptyTable:
    def test_empty_table(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "empty.xlsx"
        exporter.export_table(TABLE_DATA_FRESHNESS, out)

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "No data"


class TestHeaderStyling:
    def test_header_styling(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "styled.xlsx"
        exporter.export_table("daily_summary", out)

        wb = load_workbook(str(out))
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "00FFFFFF"
        assert header_cell.fill.start_color.rgb == "004472C4"


class TestAutoColumnWidth:
    def test_auto_column_width(self, backend_with_data, tmp_path):
        exporter = ExcelExporter(backend_with_data)
        out = tmp_path / "widths.xlsx"
        exporter.export_table("daily_summary", out)

        wb = load_workbook(str(out))
        ws = wb.active
        col_b_width = ws.column_dimensions["B"].width
        assert col_b_width > len("request_date")


class TestBuildSelectValidation:
    """REQ-002: Table and column validation in _build_select()."""

    def test_build_select_valid_table(self, backend_with_data):
        exporter = ExcelExporter(backend_with_data)
        sql, params = exporter._build_select("daily_summary")
        assert sql == "SELECT * FROM daily_summary"
        assert params == {}

    def test_build_select_valid_table_with_filters(self, backend_with_data):
        exporter = ExcelExporter(backend_with_data)
        sql, params = exporter._build_select(
            "daily_summary", filters={"bot_name": "GPTBot"}, limit=10
        )
        assert "daily_summary" in sql
        assert "bot_name = :filter_bot_name" in sql
        assert "LIMIT 10" in sql
        assert params["filter_bot_name"] == "GPTBot"

    def test_build_select_invalid_table(self, backend_with_data):
        exporter = ExcelExporter(backend_with_data)
        with pytest.raises(ValueError, match="Invalid table name"):
            exporter._build_select("Robert'; DROP TABLE--")

    def test_build_select_injection_in_column(self, backend_with_data):
        exporter = ExcelExporter(backend_with_data)
        with pytest.raises(ValueError, match="Invalid column name"):
            exporter._build_select(
                "daily_summary",
                filters={"1=1; DROP TABLE daily_summary--": "x"},
            )

    def test_build_select_invalid_date_column(self, backend_with_data):
        exporter = ExcelExporter(backend_with_data)
        with pytest.raises(ValueError, match="Invalid date column"):
            exporter._build_select(
                "daily_summary",
                filters={
                    "__date_col": "malicious_col",
                    "__start_date": "2026-01-01",
                },
            )
